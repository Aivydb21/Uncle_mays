"""
Uncle May's Produce — North Lawndale Food Hub
Steans Family Foundation Financial Model

Generates a multi-tab Excel workbook with live formulas so reviewers can
flex any input and watch the model recalculate.

Run:  python build_model.py
Out:  uncle-mays-financial-model.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

INPUT_FONT = Font(color="0000CC", bold=False, name="Calibri", size=11)
FORMULA_FONT = Font(color="000000", bold=False, name="Calibri", size=11)
LINK_FONT = Font(color="006600", bold=False, name="Calibri", size=11)
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
SECTION_FONT = Font(color="000000", bold=True, name="Calibri", size=12)
TOTAL_FONT = Font(color="000000", bold=True, name="Calibri", size=11)

HEADER_FILL = PatternFill("solid", fgColor="1F4E2C")  # Uncle May's dark green
SECTION_FILL = PatternFill("solid", fgColor="D5E3DC")
SUBTOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")
TOTAL_FILL = PatternFill("solid", fgColor="E8E5C8")
NOTES_FILL = PatternFill("solid", fgColor="FFF8E1")

THIN = Side(border_style="thin", color="999999")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=False)


def header_row(ws, row, labels, start_col=1):
    for i, label in enumerate(labels):
        c = ws.cell(row=row, column=start_col + i, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 30


def section_row(ws, row, label, span=14, start_col=1):
    c = ws.cell(row=row, column=start_col, value=label)
    c.font = SECTION_FONT
    c.fill = SECTION_FILL
    for col in range(start_col, start_col + span):
        ws.cell(row=row, column=col).fill = SECTION_FILL
    ws.row_dimensions[row].height = 22


def input_cell(ws, row, col, value, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = INPUT_FONT
    if fmt:
        c.number_format = fmt
    c.alignment = RIGHT
    return c


def formula_cell(ws, row, col, formula, fmt=None, bold=False):
    c = ws.cell(row=row, column=col, value=formula)
    c.font = TOTAL_FONT if bold else FORMULA_FONT
    if fmt:
        c.number_format = fmt
    c.alignment = RIGHT
    return c


def label_cell(ws, row, col, text, bold=False, indent=0):
    c = ws.cell(row=row, column=col, value=("  " * indent) + text)
    c.font = TOTAL_FONT if bold else FORMULA_FONT
    c.alignment = LEFT
    return c


def total_row_format(ws, row, start_col, end_col):
    for col in range(start_col, end_col + 1):
        c = ws.cell(row=row, column=col)
        c.font = TOTAL_FONT
        c.fill = TOTAL_FILL
        c.border = BORDER


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


CURR = '"$"#,##0;[Red]("$"#,##0)'
CURR_K = '"$"#,##0,"K";[Red]("$"#,##0,"K")'
PCT = '0.0%'
PCT_INT = '0%'
NUM = '#,##0'
NUM_DEC = '#,##0.00'


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------

wb = Workbook()
wb.remove(wb.active)


# ===========================================================================
# README
# ===========================================================================

ws = wb.create_sheet("README")
set_col_widths(ws, [2, 110])

ws.cell(row=1, column=2, value="Uncle May's Produce — North Lawndale Food Hub").font = Font(bold=True, size=18, color="1F4E2C")
ws.cell(row=2, column=2, value="Steans Family Foundation Financial Model").font = Font(bold=True, size=13, color="555555")
ws.cell(row=3, column=2, value="3726 W 16th Street, Chicago, IL 60623 | 3,000 sq ft | Black-owned grocery + prepared foods").font = Font(italic=True, size=10, color="555555")
ws.cell(row=4, column=2, value="Prepared: May 2026").font = Font(italic=True, size=10, color="555555")

readme_lines = [
    ("", ""),
    ("STRUCTURE", "section"),
    ("This workbook is structured around a TWO-PHASE capital deployment:", ""),
    ("    Phase 1 (Months 1-3): Produce-box delivery launch, ~$250K capital", ""),
    ("    Phase 2 (Months 4-7+): Retail buildout and operation, ~$500K capital", ""),
    ("Phase 2 release is conditional on Phase 1 milestone gates (see 'Milestones' tab).", ""),
    ("", ""),
    ("HOW TO USE", "section"),
    ("All blue-font cells on the 'Inputs' tab are user-editable assumptions.", ""),
    ("Black-font cells are formulas. Do not edit unless you know what you're changing.", ""),
    ("Green-font cells are cross-tab references.", ""),
    ("Change any input on 'Inputs' and the entire model recalculates.", ""),
    ("", ""),
    ("TAB GUIDE", "section"),
    ("    Inputs              All assumptions, color-coded blue", ""),
    ("    Capital Stack       Sources and Uses by tranche", ""),
    ("    Phase 1 Revenue     Produce-box subscriber ramp, monthly Year 1", ""),
    ("    Phase 2 Revenue     Retail department mix, monthly Year 1 + annual Y2-5", ""),
    ("    P&L Monthly Y1      Combined monthly P&L", ""),
    ("    P&L Annual          5-year P&L, 3 scenarios (base / downside / stress)", ""),
    ("    Working Capital     Cash conversion cycle and reserve modeling", ""),
    ("    Debt & DSCR         Debt schedule and coverage ratios", ""),
    ("    Sensitivity         EBITDA sensitivity to revenue, GM, and labor", ""),
    ("    Milestones          Phase 1 to Phase 2 release gates", ""),
    ("    Impact              Steans-format community outcomes", ""),
    ("", ""),
    ("KEY ASSUMPTIONS NOTE", "section"),
    ("Equipment offset at 3726 W 16th: This model assumes ~$200K of in-place equipment ", ""),
    ("(refrigeration, walk-ins, kitchen black iron, counters, shelving) at the Muldrow space.", ""),
    ("This is the single largest economic lever and the largest model dependency.", ""),
    ("An equipment audit prior to final submission is recommended to firm the number.", ""),
    ("", ""),
    ("Revenue benchmarks: Industry comparables suggest $300 to $550 per sq ft for", ""),
    ("small-format underserved-market grocery, building over 18 to 24 months.", ""),
    ("This model targets ~$200/sf Year 1 (ramp), ~$385/sf Year 2 (stabilized),", ""),
    ("rising to ~$500/sf by Year 3. See 'Comparables Memo' for sourcing.", ""),
    ("", ""),
    ("Color key:", "section"),
    ("    Blue font   = user input (editable)", ""),
    ("    Black font  = formula", ""),
    ("    Green font  = cross-tab reference", ""),
]

r = 5
for text, kind in readme_lines:
    if kind == "section":
        c = ws.cell(row=r, column=2, value=text)
        c.font = Font(bold=True, size=12, color="1F4E2C")
        c.fill = SECTION_FILL
    else:
        ws.cell(row=r, column=2, value=text).font = Font(name="Calibri", size=11)
    r += 1


# ===========================================================================
# INPUTS
# ===========================================================================

ws = wb.create_sheet("Inputs")
set_col_widths(ws, [42, 18, 18, 60])
header_row(ws, 1, ["Assumption", "Value", "Unit", "Notes / Source"])

inputs = []  # collect cell refs by name for later
named = {}


def add_input(label, value, unit, notes, fmt=None, bold=False):
    inputs.append((label, value, unit, notes, fmt, bold))


# Project basics
add_input("PROJECT BASICS", None, None, None, None, True)
add_input("Footprint (sq ft)", 3000, "sq ft", "3726 W 16th St, Muldrow space", NUM)
add_input("Phase 1 rent (Months 1-3)", 0, "$/mo", "90-day rent-free per Muldrow MOU", CURR)
add_input("Phase 2 rent subsidized rate", 1.00, "$/sf/yr", "Subsidized for first 12 months post-Phase 2 open", '"$"0.00')
add_input("Market rent (Year 2+)", 18.00, "$/sf/yr", "West-side neighborhood retail comp band: $18-$25", '"$"0.00')
add_input("Phase 2 open month", 7, "month", "Retail floor opens Month 7", NUM)

# Phase 1 produce box
add_input("PHASE 1 — PRODUCE BOX BUSINESS", None, None, None, None, True)
add_input("Average box price", 50, "$/box", "Range $25-$80 per Anthony; $50 mid-point", CURR)
add_input("Box gross margin", 0.25, "%", "25% per Anthony", PCT)
add_input("Boxes per subscriber per month", 4, "boxes", "Weekly delivery x 4 weeks", NUM)
add_input("Delivery cost per box", 4.50, "$/box", "Variable cost (driver labor, fuel, packaging)", CURR)
add_input("One-time order revenue (% of sub revenue)", 0.30, "%", "Non-subscriber walk-up and one-off orders", PCT)
add_input("Monthly subscriber churn", 0.05, "%", "5% monthly = ~46% annual retention; conservative", PCT)

# Phase 1 subscriber ramp (Months 1-12)
add_input("New subscribers — Month 1", 25, "subs", "Community launch + Pembroke + friends/family", NUM)
add_input("New subscribers — Month 2", 35, "subs", "Word of mouth and outreach", NUM)
add_input("New subscribers — Month 3", 45, "subs", "Tranche 2 gate threshold = 100 active by end of M3", NUM)
add_input("New subscribers — Month 4", 45, "subs", None, NUM)
add_input("New subscribers — Month 5", 45, "subs", None, NUM)
add_input("New subscribers — Month 6", 45, "subs", None, NUM)
add_input("New subscribers — Month 7", 40, "subs", "Retail launch absorbs some attention", NUM)
add_input("New subscribers — Month 8", 35, "subs", None, NUM)
add_input("New subscribers — Month 9", 30, "subs", None, NUM)
add_input("New subscribers — Month 10", 25, "subs", None, NUM)
add_input("New subscribers — Month 11", 25, "subs", None, NUM)
add_input("New subscribers — Month 12", 25, "subs", "Steady-state acquisition rate", NUM)

# Phase 2 retail
add_input("PHASE 2 — RETAIL", None, None, None, None, True)
add_input("Retail revenue — Month 7 (open)", 40000, "$/mo", "Soft open, limited assortment", CURR)
add_input("Retail revenue — Month 8", 55000, "$/mo", None, CURR)
add_input("Retail revenue — Month 9", 70000, "$/mo", None, CURR)
add_input("Retail revenue — Month 10", 80000, "$/mo", None, CURR)
add_input("Retail revenue — Month 11", 85000, "$/mo", None, CURR)
add_input("Retail revenue — Month 12", 90000, "$/mo", "Year 1 exit run-rate", CURR)

# Department mix and margins
add_input("Dry grocery — % of retail mix", 0.30, "%", "Industry: 35-45%; we run leaner SKU count", PCT)
add_input("Dry grocery — gross margin", 0.26, "%", "Industry 25-28%", PCT)
add_input("Produce — % of retail mix", 0.18, "%", "Pembroke direct sourcing", PCT)
add_input("Produce — gross margin", 0.42, "%", "Industry 38-45% with direct sourcing", PCT)
add_input("Dairy — % of retail mix", 0.08, "%", None, PCT)
add_input("Dairy — gross margin", 0.30, "%", None, PCT)
add_input("Meat — % of retail mix", 0.08, "%", None, PCT)
add_input("Meat — gross margin", 0.30, "%", None, PCT)
add_input("Frozen — % of retail mix", 0.06, "%", None, PCT)
add_input("Frozen — gross margin", 0.30, "%", None, PCT)
add_input("Prepared / hot foods — % of retail mix", 0.20, "%", "Highest-margin category; differentiator", PCT)
add_input("Prepared / hot foods — gross margin", 0.50, "%", "Industry 45-60%", PCT)
add_input("Deli / bakery — % of retail mix", 0.05, "%", None, PCT)
add_input("Deli / bakery — gross margin", 0.38, "%", None, PCT)
add_input("Private label — % of retail mix", 0.05, "%", "Uncle May's branded goods", PCT)
add_input("Private label — gross margin", 0.45, "%", None, PCT)

# Year 2-5 retail growth
add_input("Retail revenue Y2 growth", 0.20, "%", "Y1 exit run-rate $90K/mo annualized + 20% growth", PCT)
add_input("Retail revenue Y3 growth", 0.15, "%", "Stabilization", PCT)
add_input("Retail revenue Y4 growth", 0.08, "%", None, PCT)
add_input("Retail revenue Y5 growth", 0.05, "%", "Mature steady-state", PCT)

# Phase 1 subscriber Year 2+ assumptions
add_input("Y2 produce subscribers (avg active)", 350, "subs", "Citywide delivery, Phase 1 hardened (avg over Y2)", NUM)
add_input("Y3 produce subscribers (avg active)", 450, "subs", None, NUM)
add_input("Y4 produce subscribers (avg active)", 500, "subs", None, NUM)
add_input("Y5 produce subscribers (avg active)", 525, "subs", "Mature steady-state", NUM)

# Operating expenses
add_input("OPERATING EXPENSES", None, None, None, None, True)
add_input("Wage rate (avg loaded)", 18.40, "$/hr", "$16/hr base + 15% taxes/benefits", '"$"0.00')
add_input("FTE hours per month", 173, "hrs", "40 hr/wk x 4.33 weeks", NUM)
add_input("Phase 1 FTE — Months 1-3", 5, "FTE", "Drivers, packers, ops manager, kitchen prep", NUM)
add_input("Phase 1 FTE — Months 4-6", 6, "FTE", None, NUM)
add_input("Phase 2 FTE — Months 7-9", 8, "FTE", "Retail launch", NUM)
add_input("Phase 2 FTE — Months 10-12", 9, "FTE", None, NUM)
add_input("Y2 FTE — average", 11, "FTE", NUM)
add_input("Y3 FTE — average", 12, "FTE", NUM)
add_input("Y4 FTE — average", 13, "FTE", NUM)
add_input("Y5 FTE — average", 14, "FTE", NUM)

add_input("Utilities (Phase 1, MOU absorbs)", 0, "$/mo", "Per Muldrow MOU during Phase 1", CURR)
add_input("Utilities (Phase 2, monthly)", 2500, "$/mo", "Electricity, water, gas, internet", CURR)
add_input("Insurance (annual)", 12000, "$/yr", "GL + property + workers comp", CURR)
add_input("POS / technology (monthly)", 600, "$/mo", "Software + hardware service", CURR)
add_input("Cleaning / supplies (monthly)", 700, "$/mo", "Consumables", CURR)
add_input("Marketing / outreach (monthly Y1)", 2000, "$/mo", "Local press, community events, social", CURR)
add_input("Vehicle / delivery overhead", 1500, "$/mo", "Vehicle lease or shared, maintenance", CURR)
add_input("Professional fees (monthly)", 800, "$/mo", "Legal, accounting, compliance", CURR)
add_input("Shrink / waste (% of retail revenue)", 0.03, "%", "Industry 2-4%; produce-heavy at upper end", PCT)

# Capital structure
add_input("CAPITAL STACK", None, None, None, None, True)
add_input("Tranche 1 (Phase 1 launch)", 250000, "$", "Phase 1 produce-box business", CURR)
add_input("Tranche 2 (Phase 2 retail)", 500000, "$", "Conditional on Phase 1 milestones", CURR)
add_input("Workforce dev grant (T1 portion)", 100000, "$", "Year 1 staffing offset, Phase 1", CURR)
add_input("Workforce dev grant (T2 portion)", 50000, "$", "Year 1 staffing offset, Phase 2", CURR)
add_input("Term debt — Phase 1 portion", 100000, "$", "Equipment + inventory; 7-yr amort, 6%", CURR)
add_input("Term debt — Phase 2 portion", 350000, "$", "Buildout + equipment + inventory; 7-yr amort, 6%", CURR)
add_input("Subordinated debt — Phase 1 reserve", 50000, "$", "Phase 1 cash buffer; 7-yr, 4% (PRI-style)", CURR)
add_input("Operating reserve (subordinated)", 100000, "$", "6-9 months Phase 2 opex; T2 only", CURR)
add_input("Line of credit facility", 100000, "$", "Revolving for working capital and CCC drag", CURR)
add_input("LoC interest rate", 0.085, "%", "Prime + 1, approximate", PCT)
add_input("Term debt interest rate", 0.06, "%", "Below-market Steans-style PRI", PCT)
add_input("Term debt amortization (years)", 7, "years", NUM)
add_input("Average LoC drawn (Y1)", 60000, "$", "CCC drag during ramp", CURR)
add_input("Average LoC drawn (Y2)", 30000, "$", "Improving as vendor terms earn out", CURR)

# Working capital
add_input("WORKING CAPITAL", None, None, None, None, True)
add_input("Days Inventory Outstanding (Y1)", 25, "days", "Independent ramp; mostly fresh", NUM)
add_input("Days Inventory Outstanding (Y2+)", 22, "days", "Mature, mix shifts to faster turn", NUM)
add_input("Days Payable Outstanding (Y1)", 12, "days", "New independent; vendors require Net 7-15", NUM)
add_input("Days Payable Outstanding (Y2+)", 25, "days", "Earned over 12-18 months", NUM)
add_input("Days Sales Outstanding", 1, "days", "Card float", NUM)
add_input("Opening assortment investment", 100000, "$", "Day-1 shelf fill; partly liquidates over Months 1-6", CURR)


# Write inputs to sheet
input_row_map = {}  # label -> row number
r = 2
for label, value, unit, notes, fmt, bold in inputs:
    if value is None and unit is None:
        # section header
        section_row(ws, r, label, span=4)
    else:
        ws.cell(row=r, column=1, value=label).font = TOTAL_FONT if bold else FORMULA_FONT
        ws.cell(row=r, column=1).alignment = LEFT
        c = ws.cell(row=r, column=2, value=value)
        c.font = INPUT_FONT
        c.alignment = RIGHT
        if fmt:
            c.number_format = fmt
        ws.cell(row=r, column=3, value=unit or "").alignment = LEFT
        ws.cell(row=r, column=3).font = Font(italic=True, size=10, color="666666")
        if notes:
            ws.cell(row=r, column=4, value=notes).alignment = LEFT
            ws.cell(row=r, column=4).font = Font(italic=True, size=10, color="666666")
        input_row_map[label] = r
    r += 1


def inp(label):
    """Return absolute reference to an input by label."""
    return f"Inputs!$B${input_row_map[label]}"


# ===========================================================================
# CAPITAL STACK
# ===========================================================================

ws = wb.create_sheet("Capital Stack")
set_col_widths(ws, [38, 16, 16, 16, 50])
header_row(ws, 1, ["Sources & Uses", "Tranche 1", "Tranche 2", "Total", "Notes / Instrument"])

stack_rows = []

stack_rows.append(("SOURCES", None, None, None, None, "section"))
stack_rows.append(("Workforce development grant",
                   f"={inp('Workforce dev grant (T1 portion)')}",
                   f"={inp('Workforce dev grant (T2 portion)')}",
                   None,
                   "Steans-aligned: NLW priority is workforce/jobs/economic opportunity",
                   "row"))
stack_rows.append(("Term debt (subordinated, below-market)",
                   f"={inp('Term debt — Phase 1 portion')}",
                   f"={inp('Term debt — Phase 2 portion')}",
                   None,
                   "PRI-style 6%, 7-year amortization",
                   "row"))
stack_rows.append(("Subordinated debt (Phase 1 reserve)",
                   f"={inp('Subordinated debt — Phase 1 reserve')}",
                   0,
                   None,
                   "Phase 1 cash buffer; below-market PRI-style",
                   "row"))
stack_rows.append(("Operating reserve (subordinated, Phase 2)",
                   0,
                   f"={inp('Operating reserve (subordinated)')}",
                   None,
                   "6-9 months Phase 2 opex; held in cash at Tranche 2 close",
                   "row"))
stack_rows.append(("Total sources (deployed at close)", None, None, None, None, "subtotal"))
stack_rows.append(("Plus: Line of credit facility (off-balance, drawable)",
                   f"={inp('Line of credit facility')}",
                   0,
                   None,
                   "Revolving facility, draws as needed; not deployed at close",
                   "row"))
stack_rows.append(("", None, None, None, None, "blank"))
stack_rows.append(("USES — TRANCHE 1 (Phase 1 produce-box launch)", None, None, None, None, "section"))
stack_rows.append(("Equipment supplements (delivery, packing, refrig top-up)", 50000, 0, None, "Subject to equipment audit", "row"))
stack_rows.append(("Initial inventory (90-day produce + dry goods + packaging)", 70000, 0, None, "Phase 1 produce + light retail prep", "row"))
stack_rows.append(("Phase 1 staffing (5-7 FTE x 3-6 months net of grant)", 35000, 0, None, "Workforce grant absorbs majority", "row"))
stack_rows.append(("Working capital reserve (Phase 1)", 30000, 0, None, "CCC drag during ramp", "row"))
stack_rows.append(("Outreach, marketing, permits, legal", 30000, 0, None, "Community launch, brand, registrations", "row"))
stack_rows.append(("Software, POS, ops infrastructure", 20000, 0, None, "Subscription mgmt, route optimization, basic POS", "row"))
stack_rows.append(("Tranche 1 contingency (6%)", 15000, 0, None, None, "row"))
stack_rows.append(("Total Tranche 1 uses", None, None, None, None, "subtotal"))
stack_rows.append(("", None, None, None, None, "blank"))
stack_rows.append(("USES — TRANCHE 2 (Phase 2 retail launch)", None, None, None, None, "section"))
stack_rows.append(("Retail buildout delta (signage, finish, partitions)", 0, 130000, None, "Existing equipment offsets vanilla-shell build", "row"))
stack_rows.append(("Equipment delta (only items not on-site)", 0, 50000, None, "Subject to equipment audit", "row"))
stack_rows.append(("Inventory expansion (full retail assortment)", 0, 100000, None, "Day-1 shelf fill", "row"))
stack_rows.append(("Phase 2 staffing scale-up (months 7-12 net of grant)", 0, 100000, None, "Retail launch labor", "row"))
stack_rows.append(("Operating reserve (6-9 months Phase 2 opex)", 0, 100000, None, "Critical line; absent in prior model versions", "row"))
stack_rows.append(("Phase 2 contingency (4%)", 0, 20000, None, None, "row"))
stack_rows.append(("Total Tranche 2 uses", None, None, None, None, "subtotal"))
stack_rows.append(("", None, None, None, None, "blank"))
stack_rows.append(("Total uses", None, None, None, None, "total"))
stack_rows.append(("", None, None, None, None, "blank"))
stack_rows.append(("Sources less uses (must be zero)", None, None, None, None, "check"))

r = 2
sources_start = None
sources_end = None
t1_use_start = None
t1_use_end = None
t2_use_start = None
t2_use_end = None

for entry in stack_rows:
    label, t1, t2, total, notes, kind = entry
    if kind == "section":
        section_row(ws, r, label, span=5)
    elif kind == "blank":
        pass
    elif kind == "subtotal":
        if "Tranche 1 uses" in label:
            t1_use_end = r - 1
            ws.cell(row=r, column=1, value="    " + label).font = TOTAL_FONT
            formula_cell(ws, r, 2, f"=SUM(B{t1_use_start}:B{t1_use_end})", CURR, bold=True)
            formula_cell(ws, r, 3, "=0", CURR, bold=True)
            formula_cell(ws, r, 4, f"=B{r}+C{r}", CURR, bold=True)
            total_row_format(ws, r, 1, 5)
        elif "Tranche 2 uses" in label:
            t2_use_end = r - 1
            ws.cell(row=r, column=1, value="    " + label).font = TOTAL_FONT
            formula_cell(ws, r, 2, "=0", CURR, bold=True)
            formula_cell(ws, r, 3, f"=SUM(C{t2_use_start}:C{t2_use_end})", CURR, bold=True)
            formula_cell(ws, r, 4, f"=B{r}+C{r}", CURR, bold=True)
            total_row_format(ws, r, 1, 5)
        else:  # Total sources
            sources_end = r - 1
            ws.cell(row=r, column=1, value=label).font = TOTAL_FONT
            formula_cell(ws, r, 2, f"=SUM(B{sources_start}:B{sources_end})", CURR, bold=True)
            formula_cell(ws, r, 3, f"=SUM(C{sources_start}:C{sources_end})", CURR, bold=True)
            formula_cell(ws, r, 4, f"=B{r}+C{r}", CURR, bold=True)
            total_row_format(ws, r, 1, 5)
    elif kind == "total":
        ws.cell(row=r, column=1, value=label).font = TOTAL_FONT
        # Sum the two subtotal rows
        formula_cell(ws, r, 2, f"=B{t1_use_end+1}+B{t2_use_end+1}", CURR, bold=True)
        formula_cell(ws, r, 3, f"=C{t1_use_end+1}+C{t2_use_end+1}", CURR, bold=True)
        formula_cell(ws, r, 4, f"=B{r}+C{r}", CURR, bold=True)
        total_row_format(ws, r, 1, 5)
    elif kind == "check":
        ws.cell(row=r, column=1, value=label).font = TOTAL_FONT
        # sources_total minus uses_total — note offset
        # Find the source-total row and uses-total row dynamically
        # sources total is at row that was the "Total sources" row; let's hold it
        sources_total_row = sources_end + 1  # approximation - we'll fix below
        # Easier: just sum specific rows known to be totals
        formula_cell(ws, r, 4, f"=D{sources_total_row}-D{r-2}", CURR, bold=True)
    else:  # data row
        if label.startswith("    ") or sources_start is None and "SOURCES" not in label:
            pass
        if sources_start is None:
            sources_start = r
        ws.cell(row=r, column=1, value=label).font = FORMULA_FONT
        ws.cell(row=r, column=1).alignment = LEFT
        if isinstance(t1, str):
            formula_cell(ws, r, 2, t1, CURR)
        else:
            input_cell(ws, r, 2, t1, CURR)
        if isinstance(t2, str):
            formula_cell(ws, r, 3, t2, CURR)
        else:
            input_cell(ws, r, 3, t2, CURR)
        formula_cell(ws, r, 4, f"=B{r}+C{r}", CURR)
        if notes:
            ws.cell(row=r, column=5, value=notes).font = Font(italic=True, size=10, color="666666")
            ws.cell(row=r, column=5).alignment = LEFT
        # Track use-section starts
        if t1_use_start is None and "Equipment supplements" in label:
            t1_use_start = r
        if t2_use_start is None and "Retail buildout delta" in label:
            t2_use_start = r
    r += 1


# ===========================================================================
# PHASE 1 REVENUE — produce box, monthly Year 1
# ===========================================================================

ws = wb.create_sheet("Phase 1 Revenue")
set_col_widths(ws, [36] + [12]*13)
months = [f"M{i}" for i in range(1, 13)]
header_row(ws, 1, ["Phase 1 — Produce Box Business"] + months + ["Year 1 Total"])

# Row layout
# Row 2: New subscribers
label_cell(ws, 2, 1, "New subscribers (gross adds)")
for i, m in enumerate(range(1, 13)):
    formula_cell(ws, 2, 1 + i + 1, f"={inp(f'New subscribers — Month {m}')}", NUM)
formula_cell(ws, 2, 14, "=SUM(B2:M2)", NUM)

# Row 3: Churn
label_cell(ws, 3, 1, "Subscribers churned")
formula_cell(ws, 3, 2, "=0", NUM)  # M1 nobody to churn
for i in range(2, 14):  # M2 to M12
    col = i + 1
    prev_col = get_column_letter(col - 1)
    # Active prior month (row 4) * churn rate
    formula_cell(ws, 3, col, f"=ROUND({prev_col}4*{inp('Monthly subscriber churn')},0)", NUM)
formula_cell(ws, 3, 14, "=SUM(B3:M3)", NUM)

# Row 4: Active subscribers (end of month)
label_cell(ws, 4, 1, "Active subscribers (end of month)", bold=True)
formula_cell(ws, 4, 2, "=B2-B3", NUM, bold=True)
for i in range(2, 13):
    col = i + 1
    prev_col = get_column_letter(col - 1)
    cur_col = get_column_letter(col)
    formula_cell(ws, 4, col, f"={prev_col}4+{cur_col}2-{cur_col}3", NUM, bold=True)
total_row_format(ws, 4, 1, 14)
formula_cell(ws, 4, 14, "=M4", NUM, bold=True)  # ending active

# Row 5: Avg active subscribers (for revenue calc)
label_cell(ws, 5, 1, "  Avg active subscribers (this month)")
formula_cell(ws, 5, 2, "=B4/2", NUM)
for i in range(2, 13):
    col = i + 1
    prev_col = get_column_letter(col - 1)
    cur_col = get_column_letter(col)
    formula_cell(ws, 5, col, f"=({prev_col}4+{cur_col}4)/2", NUM)

# Row 6: Boxes shipped per month
label_cell(ws, 6, 1, "Boxes shipped (subscriber + one-time)")
for col in range(2, 14):
    cur_col = get_column_letter(col)
    # subscribers * boxes per sub per month, plus one-time as % of subs revenue (proxy)
    formula_cell(ws, 6, col, f"=ROUND({cur_col}5*{inp('Boxes per subscriber per month')}*(1+{inp('One-time order revenue (% of sub revenue)')}),0)", NUM)
formula_cell(ws, 6, 14, "=SUM(B6:M6)", NUM)

# Row 7: Subscription revenue
label_cell(ws, 7, 1, "Subscription revenue")
for col in range(2, 14):
    cur_col = get_column_letter(col)
    formula_cell(ws, 7, col, f"={cur_col}5*{inp('Boxes per subscriber per month')}*{inp('Average box price')}", CURR)
formula_cell(ws, 7, 14, "=SUM(B7:M7)", CURR)

# Row 8: One-time order revenue
label_cell(ws, 8, 1, "One-time order revenue")
for col in range(2, 14):
    cur_col = get_column_letter(col)
    formula_cell(ws, 8, col, f"={cur_col}7*{inp('One-time order revenue (% of sub revenue)')}", CURR)
formula_cell(ws, 8, 14, "=SUM(B8:M8)", CURR)

# Row 9: Total Phase 1 revenue
label_cell(ws, 9, 1, "Total Phase 1 revenue", bold=True)
for col in range(2, 14):
    cur_col = get_column_letter(col)
    formula_cell(ws, 9, col, f"={cur_col}7+{cur_col}8", CURR, bold=True)
formula_cell(ws, 9, 14, "=SUM(B9:M9)", CURR, bold=True)
total_row_format(ws, 9, 1, 14)

# Row 10: COGS
label_cell(ws, 10, 1, "COGS (75% of revenue)")
for col in range(2, 14):
    cur_col = get_column_letter(col)
    formula_cell(ws, 10, col, f"={cur_col}9*(1-{inp('Box gross margin')})", CURR)
formula_cell(ws, 10, 14, "=SUM(B10:M10)", CURR)

# Row 11: Gross profit
label_cell(ws, 11, 1, "Gross profit", bold=True)
for col in range(2, 14):
    cur_col = get_column_letter(col)
    formula_cell(ws, 11, col, f"={cur_col}9-{cur_col}10", CURR, bold=True)
formula_cell(ws, 11, 14, "=SUM(B11:M11)", CURR, bold=True)
total_row_format(ws, 11, 1, 14)

# Row 12: Delivery cost (variable, separate from COGS)
label_cell(ws, 12, 1, "Delivery cost (driver, fuel, packaging)")
for col in range(2, 14):
    cur_col = get_column_letter(col)
    formula_cell(ws, 12, col, f"={cur_col}6*{inp('Delivery cost per box')}", CURR)
formula_cell(ws, 12, 14, "=SUM(B12:M12)", CURR)

# Row 13: Contribution margin (gross profit less delivery)
label_cell(ws, 13, 1, "Contribution margin (Phase 1)", bold=True)
for col in range(2, 14):
    cur_col = get_column_letter(col)
    formula_cell(ws, 13, col, f"={cur_col}11-{cur_col}12", CURR, bold=True)
formula_cell(ws, 13, 14, "=SUM(B13:M13)", CURR, bold=True)
total_row_format(ws, 13, 1, 14)

# Tranche 2 gate check
section_row(ws, 15, "Tranche 2 release gate check (end of Month 3)", span=14)
label_cell(ws, 16, 1, "Active subscribers at M3 (target: 100+)")
formula_cell(ws, 16, 2, "=D4", NUM, bold=True)
label_cell(ws, 17, 1, "M3 monthly revenue (target: $15K-$25K)")
formula_cell(ws, 17, 2, "=D9", CURR, bold=True)


# ===========================================================================
# PHASE 2 REVENUE — retail
# ===========================================================================

ws = wb.create_sheet("Phase 2 Revenue")
set_col_widths(ws, [36] + [12]*13)
header_row(ws, 1, ["Phase 2 — Retail Floor"] + months + ["Year 1 Total"])

# Months 1-6 are zero, Months 7-12 from inputs
label_cell(ws, 2, 1, "Total retail revenue", bold=True)
for col in range(2, 8):  # M1 to M6
    formula_cell(ws, 2, col, "=0", CURR, bold=True)
month_inputs = [
    "Retail revenue — Month 7 (open)",
    "Retail revenue — Month 8",
    "Retail revenue — Month 9",
    "Retail revenue — Month 10",
    "Retail revenue — Month 11",
    "Retail revenue — Month 12",
]
for i, label in enumerate(month_inputs):
    col = 8 + i  # M7 starts at col 8
    formula_cell(ws, 2, col, f"={inp(label)}", CURR, bold=True)
formula_cell(ws, 2, 14, "=SUM(B2:M2)", CURR, bold=True)
total_row_format(ws, 2, 1, 14)

# Department breakdown (mix x revenue)
section_row(ws, 4, "Revenue by department (= mix % * total revenue)", span=14)
depts = [
    ("Dry grocery",        "Dry grocery — % of retail mix",        "Dry grocery — gross margin"),
    ("Produce",            "Produce — % of retail mix",            "Produce — gross margin"),
    ("Dairy",              "Dairy — % of retail mix",              "Dairy — gross margin"),
    ("Meat",               "Meat — % of retail mix",               "Meat — gross margin"),
    ("Frozen",             "Frozen — % of retail mix",             "Frozen — gross margin"),
    ("Prepared / hot foods", "Prepared / hot foods — % of retail mix", "Prepared / hot foods — gross margin"),
    ("Deli / bakery",      "Deli / bakery — % of retail mix",      "Deli / bakery — gross margin"),
    ("Private label",      "Private label — % of retail mix",      "Private label — gross margin"),
]
dept_start = 5
for i, (name, mix_lbl, gm_lbl) in enumerate(depts):
    r = dept_start + i
    label_cell(ws, r, 1, f"  {name} ({mix_lbl.split(' — ')[0]} mix)")
    for col in range(2, 14):
        cur_col = get_column_letter(col)
        formula_cell(ws, r, col, f"={cur_col}2*{inp(mix_lbl)}", CURR)

# Mix check row
mix_check_row = dept_start + len(depts)
label_cell(ws, mix_check_row, 1, "  Mix check (=100%)")
mix_sum = "+".join(inp(m) for _, m, _ in depts)
formula_cell(ws, mix_check_row, 2, f"={mix_sum}", PCT)

# COGS by department
section_row(ws, mix_check_row + 2, "COGS by department (= revenue * (1-GM))", span=14)
cogs_start = mix_check_row + 3
for i, (name, mix_lbl, gm_lbl) in enumerate(depts):
    r = cogs_start + i
    label_cell(ws, r, 1, f"  {name} COGS")
    rev_row = dept_start + i
    for col in range(2, 14):
        cur_col = get_column_letter(col)
        formula_cell(ws, r, col, f"={cur_col}{rev_row}*(1-{inp(gm_lbl)})", CURR)

# Total COGS
total_cogs_row = cogs_start + len(depts)
label_cell(ws, total_cogs_row, 1, "Total retail COGS", bold=True)
for col in range(2, 14):
    cur_col = get_column_letter(col)
    formula_cell(ws, total_cogs_row, col, f"=SUM({cur_col}{cogs_start}:{cur_col}{cogs_start+len(depts)-1})", CURR, bold=True)
formula_cell(ws, total_cogs_row, 14, f"=SUM(B{total_cogs_row}:M{total_cogs_row})", CURR, bold=True)
total_row_format(ws, total_cogs_row, 1, 14)

# Gross profit
gp_row = total_cogs_row + 1
label_cell(ws, gp_row, 1, "Total retail gross profit", bold=True)
for col in range(2, 14):
    cur_col = get_column_letter(col)
    formula_cell(ws, gp_row, col, f"={cur_col}2-{cur_col}{total_cogs_row}", CURR, bold=True)
formula_cell(ws, gp_row, 14, f"=SUM(B{gp_row}:M{gp_row})", CURR, bold=True)
total_row_format(ws, gp_row, 1, 14)

# Blended GM
gm_row = gp_row + 1
label_cell(ws, gm_row, 1, "Blended retail GM %")
for col in range(2, 14):
    cur_col = get_column_letter(col)
    formula_cell(ws, gm_row, col, f"=IFERROR({cur_col}{gp_row}/{cur_col}2,0)", PCT)
formula_cell(ws, gm_row, 14, f"=IFERROR(N{gp_row}/N2,0)", PCT)


# ===========================================================================
# P&L MONTHLY YEAR 1
# ===========================================================================

ws = wb.create_sheet("P&L Monthly Y1")
set_col_widths(ws, [36] + [12]*13)
header_row(ws, 1, ["P&L (Combined Phase 1 + Phase 2)"] + months + ["Year 1 Total"])

# Revenue
section_row(ws, 2, "Revenue", span=14)
label_cell(ws, 3, 1, "Phase 1 produce-box revenue")
for col in range(2, 15):
    cur_col = get_column_letter(col)
    target_col = "N" if col == 14 else cur_col
    formula_cell(ws, 3, col, f"='Phase 1 Revenue'!{target_col}9", CURR)
    ws.cell(row=3, column=col).font = LINK_FONT

label_cell(ws, 4, 1, "Phase 2 retail revenue")
for col in range(2, 15):
    cur_col = get_column_letter(col)
    target_col = "N" if col == 14 else cur_col
    formula_cell(ws, 4, col, f"='Phase 2 Revenue'!{target_col}2", CURR)
    ws.cell(row=4, column=col).font = LINK_FONT

label_cell(ws, 5, 1, "Total revenue", bold=True)
for col in range(2, 15):
    cur_col = get_column_letter(col)
    formula_cell(ws, 5, col, f"={cur_col}3+{cur_col}4", CURR, bold=True)
total_row_format(ws, 5, 1, 14)

# COGS
section_row(ws, 7, "Cost of goods sold", span=14)
label_cell(ws, 8, 1, "Phase 1 COGS")
for col in range(2, 15):
    cur_col = get_column_letter(col)
    target_col = "N" if col == 14 else cur_col
    formula_cell(ws, 8, col, f"='Phase 1 Revenue'!{target_col}10", CURR)
    ws.cell(row=8, column=col).font = LINK_FONT

label_cell(ws, 9, 1, "Phase 2 COGS")
for col in range(2, 15):
    cur_col = get_column_letter(col)
    target_col = "N" if col == 14 else cur_col
    # cogs row in phase 2 sheet
    p2_cogs_row = total_cogs_row
    formula_cell(ws, 9, col, f"='Phase 2 Revenue'!{target_col}{p2_cogs_row}", CURR)
    ws.cell(row=9, column=col).font = LINK_FONT

label_cell(ws, 10, 1, "Total COGS", bold=True)
for col in range(2, 15):
    cur_col = get_column_letter(col)
    formula_cell(ws, 10, col, f"={cur_col}8+{cur_col}9", CURR, bold=True)
total_row_format(ws, 10, 1, 14)

# Gross profit
label_cell(ws, 11, 1, "Gross profit", bold=True)
for col in range(2, 15):
    cur_col = get_column_letter(col)
    formula_cell(ws, 11, col, f"={cur_col}5-{cur_col}10", CURR, bold=True)
total_row_format(ws, 11, 1, 14)

label_cell(ws, 12, 1, "  Gross margin %")
for col in range(2, 15):
    cur_col = get_column_letter(col)
    formula_cell(ws, 12, col, f"=IFERROR({cur_col}11/{cur_col}5,0)", PCT)

# Operating expenses
section_row(ws, 14, "Operating expenses", span=14)

# Labor
label_cell(ws, 15, 1, "  Labor (gross)")
fte_per_month = [
    inp("Phase 1 FTE — Months 1-3"), inp("Phase 1 FTE — Months 1-3"), inp("Phase 1 FTE — Months 1-3"),
    inp("Phase 1 FTE — Months 4-6"), inp("Phase 1 FTE — Months 4-6"), inp("Phase 1 FTE — Months 4-6"),
    inp("Phase 2 FTE — Months 7-9"), inp("Phase 2 FTE — Months 7-9"), inp("Phase 2 FTE — Months 7-9"),
    inp("Phase 2 FTE — Months 10-12"), inp("Phase 2 FTE — Months 10-12"), inp("Phase 2 FTE — Months 10-12"),
]
for i in range(12):
    col = 2 + i
    formula_cell(ws, 15, col, f"={fte_per_month[i]}*{inp('Wage rate (avg loaded)')}*{inp('FTE hours per month')}", CURR)
formula_cell(ws, 15, 14, "=SUM(B15:M15)", CURR)

# Workforce grant offset
label_cell(ws, 16, 1, "  Less: Workforce dev grant offset (amortized)")
# Amortize $150K total over 12 months = $12.5K/mo
formula_cell(ws, 16, 14, f"=-({inp('Workforce dev grant (T1 portion)')}+{inp('Workforce dev grant (T2 portion)')})", CURR)
for col in range(2, 14):
    formula_cell(ws, 16, col, f"=$N16/12", CURR)

# Net labor
label_cell(ws, 17, 1, "  Net labor")
for col in range(2, 15):
    cur_col = get_column_letter(col)
    formula_cell(ws, 17, col, f"={cur_col}15+{cur_col}16", CURR)

# Rent
label_cell(ws, 18, 1, "  Rent (subsidized)")
for i, col in enumerate(range(2, 14)):
    if i < 6:  # M1-M6 free
        formula_cell(ws, 18, col, f"={inp('Phase 1 rent (Months 1-3)')}", CURR)
    else:  # M7-M12 subsidized $1/sf/yr
        formula_cell(ws, 18, col, f"={inp('Phase 2 rent subsidized rate')}*{inp('Footprint (sq ft)')}/12", CURR)
formula_cell(ws, 18, 14, "=SUM(B18:M18)", CURR)

# Utilities
label_cell(ws, 19, 1, "  Utilities")
for col in range(2, 8):
    formula_cell(ws, 19, col, f"={inp('Utilities (Phase 1, MOU absorbs)')}", CURR)
for col in range(8, 14):
    formula_cell(ws, 19, col, f"={inp('Utilities (Phase 2, monthly)')}", CURR)
formula_cell(ws, 19, 14, "=SUM(B19:M19)", CURR)

# Insurance
label_cell(ws, 20, 1, "  Insurance")
for col in range(2, 14):
    formula_cell(ws, 20, col, f"={inp('Insurance (annual)')}/12", CURR)
formula_cell(ws, 20, 14, "=SUM(B20:M20)", CURR)

# POS
label_cell(ws, 21, 1, "  POS / technology")
for col in range(2, 14):
    formula_cell(ws, 21, col, f"={inp('POS / technology (monthly)')}", CURR)
formula_cell(ws, 21, 14, "=SUM(B21:M21)", CURR)

# Cleaning
label_cell(ws, 22, 1, "  Cleaning / supplies")
for col in range(2, 14):
    formula_cell(ws, 22, col, f"={inp('Cleaning / supplies (monthly)')}", CURR)
formula_cell(ws, 22, 14, "=SUM(B22:M22)", CURR)

# Marketing
label_cell(ws, 23, 1, "  Marketing / outreach")
for col in range(2, 14):
    formula_cell(ws, 23, col, f"={inp('Marketing / outreach (monthly Y1)')}", CURR)
formula_cell(ws, 23, 14, "=SUM(B23:M23)", CURR)

# Vehicle / delivery
label_cell(ws, 24, 1, "  Vehicle / delivery overhead")
for col in range(2, 14):
    formula_cell(ws, 24, col, f"={inp('Vehicle / delivery overhead')}", CURR)
formula_cell(ws, 24, 14, "=SUM(B24:M24)", CURR)

# Delivery variable cost (already in P1 calc; treat separately if needed)
label_cell(ws, 25, 1, "  Phase 1 delivery variable cost")
for col in range(2, 15):
    cur_col = get_column_letter(col)
    target_col = "N" if col == 14 else cur_col
    formula_cell(ws, 25, col, f"='Phase 1 Revenue'!{target_col}12", CURR)
    ws.cell(row=25, column=col).font = LINK_FONT

# Professional fees
label_cell(ws, 26, 1, "  Professional fees")
for col in range(2, 14):
    formula_cell(ws, 26, col, f"={inp('Professional fees (monthly)')}", CURR)
formula_cell(ws, 26, 14, "=SUM(B26:M26)", CURR)

# Shrink (Phase 2 only)
label_cell(ws, 27, 1, "  Shrink / waste (% of retail revenue)")
for col in range(2, 15):
    cur_col = get_column_letter(col)
    formula_cell(ws, 27, col, f"={cur_col}4*{inp('Shrink / waste (% of retail revenue)')}", CURR)

# Total opex
label_cell(ws, 28, 1, "Total operating expenses", bold=True)
for col in range(2, 15):
    cur_col = get_column_letter(col)
    formula_cell(ws, 28, col, f"=SUM({cur_col}17:{cur_col}27)", CURR, bold=True)
total_row_format(ws, 28, 1, 14)

# EBITDA
label_cell(ws, 30, 1, "EBITDA", bold=True)
for col in range(2, 15):
    cur_col = get_column_letter(col)
    formula_cell(ws, 30, col, f"={cur_col}11-{cur_col}28", CURR, bold=True)
total_row_format(ws, 30, 1, 14)

label_cell(ws, 31, 1, "  EBITDA margin %")
for col in range(2, 15):
    cur_col = get_column_letter(col)
    formula_cell(ws, 31, col, f"=IFERROR({cur_col}30/{cur_col}5,0)", PCT)

# Below the line
section_row(ws, 33, "Below EBITDA", span=14)
label_cell(ws, 34, 1, "  D&A (estimated, equipment 7-yr SL)")
# Total equipment ~ $50K T1 + $50K T2 = $100K / 7 yr / 12 = $1190/mo, conservatively
for col in range(2, 14):
    formula_cell(ws, 34, col, "=1200", CURR)
formula_cell(ws, 34, 14, "=SUM(B34:M34)", CURR)

label_cell(ws, 35, 1, "  Interest — term debt")
# Simple: outstanding * rate / 12, full balance for simplicity Y1
for col in range(2, 14):
    formula_cell(ws, 35, col, f"=({inp('Term debt — Phase 1 portion')}+{inp('Term debt — Phase 2 portion')})*{inp('Term debt interest rate')}/12", CURR)
formula_cell(ws, 35, 14, "=SUM(B35:M35)", CURR)

label_cell(ws, 36, 1, "  Interest — line of credit")
for col in range(2, 14):
    formula_cell(ws, 36, col, f"={inp('Average LoC drawn (Y1)')}*{inp('LoC interest rate')}/12", CURR)
formula_cell(ws, 36, 14, "=SUM(B36:M36)", CURR)

label_cell(ws, 38, 1, "Net income (pre-tax)", bold=True)
for col in range(2, 15):
    cur_col = get_column_letter(col)
    formula_cell(ws, 38, col, f"={cur_col}30-{cur_col}34-{cur_col}35-{cur_col}36", CURR, bold=True)
total_row_format(ws, 38, 1, 14)


# ===========================================================================
# P&L ANNUAL Y1-Y5 with three scenarios
# ===========================================================================

ws = wb.create_sheet("P&L Annual")
set_col_widths(ws, [38] + [13]*15 + [3])
# 5 columns (Y1-Y5) for Base, then 5 for Downside, then 5 for Stress
# That's 15 data columns + 1 label column

# Build header rows
# Row 1: scenario headers (merged)
ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=6)
c = ws.cell(row=1, column=2, value="BASE CASE")
c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER
ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=11)
c = ws.cell(row=1, column=7, value="DOWNSIDE")
c.font = HEADER_FONT; c.fill = PatternFill("solid", fgColor="A04040"); c.alignment = CENTER
ws.merge_cells(start_row=1, start_column=12, end_row=1, end_column=16)
c = ws.cell(row=1, column=12, value="STRESS")
c.font = HEADER_FONT; c.fill = PatternFill("solid", fgColor="6B2020"); c.alignment = CENTER

# Row 2: Y1-Y5 sub-headers
ws.cell(row=2, column=1, value="Annual P&L").font = HEADER_FONT
ws.cell(row=2, column=1).fill = HEADER_FILL
ws.cell(row=2, column=1).alignment = CENTER
years = ["Y1", "Y2", "Y3", "Y4", "Y5"]
for scenario_idx, scenario in enumerate(["base", "down", "stress"]):
    for i, y in enumerate(years):
        col = 2 + scenario_idx * 5 + i
        c = ws.cell(row=2, column=col, value=y)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL if scenario == "base" else (
            PatternFill("solid", fgColor="A04040") if scenario == "down"
            else PatternFill("solid", fgColor="6B2020"))
        c.alignment = CENTER
ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 22

# Define scenario adjustments
# Base: revenue and GM as input
# Downside: revenue x 0.85, GM -3pts
# Stress: revenue x 0.70, GM -5pts, labor +2pts of revenue

# Year 1 base values pulled from monthly P&L
# Year 2-5 driven by growth rates

# Labels and Y1 base values
section_row(ws, 4, "Revenue", span=16)

# Helper: produce-box subscriber revenue, by year (annual)
# Y1 from Phase 1 sheet N9
# Y2-Y5 from inputs (avg subscribers x box price x 4 boxes/mo x 12 months) x (1 + one-time%)

# Phase 1 produce revenue annual
label_cell(ws, 5, 1, "  Phase 1 produce-box revenue (base)")
formula_cell(ws, 5, 2, "='Phase 1 Revenue'!N9", CURR)
ws.cell(row=5, column=2).font = LINK_FONT
y2_lbl = "Y2 produce subscribers (avg active)"
y3_lbl = "Y3 produce subscribers (avg active)"
y4_lbl = "Y4 produce subscribers (avg active)"
y5_lbl = "Y5 produce subscribers (avg active)"
y_subs = [y2_lbl, y3_lbl, y4_lbl, y5_lbl]
for i, lbl in enumerate(y_subs):
    col = 3 + i  # C-F
    formula_cell(ws, 5, col, f"={inp(lbl)}*{inp('Boxes per subscriber per month')}*{inp('Average box price')}*12*(1+{inp('One-time order revenue (% of sub revenue)')})", CURR)

# Phase 2 retail revenue annual
label_cell(ws, 6, 1, "  Phase 2 retail revenue (base)")
formula_cell(ws, 6, 2, "='Phase 2 Revenue'!N2", CURR)
ws.cell(row=6, column=2).font = LINK_FONT
# Y2: annualize Y1 EXIT run-rate (M12 retail revenue * 12), then apply growth
formula_cell(ws, 6, 3, f"='Phase 2 Revenue'!M2*12*(1+{inp('Retail revenue Y2 growth')})", CURR)
formula_cell(ws, 6, 4, f"=C6*(1+{inp('Retail revenue Y3 growth')})", CURR)
formula_cell(ws, 6, 5, f"=D6*(1+{inp('Retail revenue Y4 growth')})", CURR)
formula_cell(ws, 6, 6, f"=E6*(1+{inp('Retail revenue Y5 growth')})", CURR)

# Total revenue base
label_cell(ws, 7, 1, "Total revenue (base)", bold=True)
for col in range(2, 7):
    cur_col = get_column_letter(col)
    formula_cell(ws, 7, col, f"={cur_col}5+{cur_col}6", CURR, bold=True)
total_row_format(ws, 7, 1, 6)

# Downside (col 7-11): revenue x 0.85
label_cell(ws, 7, 7, "")  # spacer
for i in range(5):
    col = 7 + i
    base_col = get_column_letter(2 + i)
    formula_cell(ws, 7, col, f"={base_col}7*0.85", CURR, bold=True)

# Stress (col 12-16): revenue x 0.70
for i in range(5):
    col = 12 + i
    base_col = get_column_letter(2 + i)
    formula_cell(ws, 7, col, f"={base_col}7*0.70", CURR, bold=True)
total_row_format(ws, 7, 7, 16)

# Gross margin assumption per scenario
section_row(ws, 9, "Gross margin %", span=16)
label_cell(ws, 10, 1, "Blended GM %")
# Year 1 base = from Y1 P&L
formula_cell(ws, 10, 2, "='P&L Monthly Y1'!N12", PCT)
ws.cell(row=10, column=2).font = LINK_FONT
# Y2-Y5 base GM ramp (Y1 computed from monthly P&L)
gm_ramp_base = [0.33, 0.35, 0.36, 0.37]
for i, gm in enumerate(gm_ramp_base):
    col = 3 + i
    input_cell(ws, 10, col, gm, PCT)
# Downside: -3pts
for i in range(5):
    col = 7 + i
    base_col = get_column_letter(2 + i)
    formula_cell(ws, 10, col, f"={base_col}10-0.03", PCT)
# Stress: -5pts
for i in range(5):
    col = 12 + i
    base_col = get_column_letter(2 + i)
    formula_cell(ws, 10, col, f"={base_col}10-0.05", PCT)

# COGS
label_cell(ws, 11, 1, "  COGS")
for col in range(2, 17):
    cur_col = get_column_letter(col)
    formula_cell(ws, 11, col, f"={cur_col}7*(1-{cur_col}10)", CURR)

# Gross profit
label_cell(ws, 12, 1, "Gross profit", bold=True)
for col in range(2, 17):
    cur_col = get_column_letter(col)
    formula_cell(ws, 12, col, f"={cur_col}7-{cur_col}11", CURR, bold=True)
total_row_format(ws, 12, 1, 16)

# Opex (% of revenue) — base assumption
section_row(ws, 14, "Operating expenses", span=16)

# Labor as % of revenue
label_cell(ws, 15, 1, "  Labor (loaded, post-grant Y1)")
# Y1 base from monthly
formula_cell(ws, 15, 2, "='P&L Monthly Y1'!N17", CURR)
ws.cell(row=15, column=2).font = LINK_FONT
# Y2-Y5: FTE * wage * hours * 12
y_fte_lbls = ["Y2 FTE — average", "Y3 FTE — average", "Y4 FTE — average", "Y5 FTE — average"]
for i, lbl in enumerate(y_fte_lbls):
    col = 3 + i
    formula_cell(ws, 15, col, f"={inp(lbl)}*{inp('Wage rate (avg loaded)')}*{inp('FTE hours per month')}*12", CURR)
# Downside same labor (sticky)
for i in range(5):
    col = 7 + i
    base_col = get_column_letter(2 + i)
    formula_cell(ws, 15, col, f"={base_col}15", CURR)
# Stress: labor + 2% of revenue
for i in range(5):
    col = 12 + i
    base_col = get_column_letter(2 + i)
    formula_cell(ws, 15, col, f"={base_col}15+{get_column_letter(12+i)}7*0.02", CURR)

# Other opex (rent, utilities, insurance, etc.) — annual
label_cell(ws, 16, 1, "  Rent")
formula_cell(ws, 16, 2, f"='P&L Monthly Y1'!N18", CURR)
ws.cell(row=16, column=2).font = LINK_FONT
# Y2 onward at market rate (with subsidy decreasing if any)
formula_cell(ws, 16, 3, f"={inp('Market rent (Year 2+)')}*{inp('Footprint (sq ft)')}", CURR)
for i in range(2, 5):
    col = 3 + i
    base_col = get_column_letter(2 + i)
    formula_cell(ws, 16, col, f"=C16*1.03^({i-1})", CURR)  # 3% rent escalator
# Downside / Stress: same rent
for i in range(5):
    formula_cell(ws, 16, 7 + i, f"={get_column_letter(2+i)}16", CURR)
    formula_cell(ws, 16, 12 + i, f"={get_column_letter(2+i)}16", CURR)

label_cell(ws, 17, 1, "  Utilities, insurance, POS, supplies, prof fees, vehicle, marketing")
# Y1 base = sum of those rows from monthly P&L
formula_cell(ws, 17, 2, "='P&L Monthly Y1'!N19+'P&L Monthly Y1'!N20+'P&L Monthly Y1'!N21+'P&L Monthly Y1'!N22+'P&L Monthly Y1'!N23+'P&L Monthly Y1'!N24+'P&L Monthly Y1'!N25+'P&L Monthly Y1'!N26", CURR)
ws.cell(row=17, column=2).font = LINK_FONT
# Y2+: 4% annual escalator
for i in range(1, 5):
    col = 2 + i
    prev_col = get_column_letter(col - 1)
    formula_cell(ws, 17, col, f"={prev_col}17*1.04", CURR)
# Down/Stress same
for i in range(5):
    formula_cell(ws, 17, 7 + i, f"={get_column_letter(2+i)}17", CURR)
    formula_cell(ws, 17, 12 + i, f"={get_column_letter(2+i)}17", CURR)

label_cell(ws, 18, 1, "  Shrink / waste")
for i in range(5):
    col = 2 + i
    cur_col = get_column_letter(col)
    formula_cell(ws, 18, col, f"={cur_col}6*{inp('Shrink / waste (% of retail revenue)')}", CURR)
for i in range(5):
    col = 7 + i
    cur_col = get_column_letter(col)
    base_col = get_column_letter(2 + i)
    # Shrink scales with downside revenue
    formula_cell(ws, 18, col, f"={cur_col}7*{inp('Shrink / waste (% of retail revenue)')}*0.6", CURR)  # less retail revenue
for i in range(5):
    col = 12 + i
    cur_col = get_column_letter(col)
    formula_cell(ws, 18, col, f"={cur_col}7*({inp('Shrink / waste (% of retail revenue)')}+0.01)", CURR)

# Total opex
label_cell(ws, 19, 1, "Total opex", bold=True)
for col in range(2, 17):
    cur_col = get_column_letter(col)
    formula_cell(ws, 19, col, f"={cur_col}15+{cur_col}16+{cur_col}17+{cur_col}18", CURR, bold=True)
total_row_format(ws, 19, 1, 16)

# EBITDA
label_cell(ws, 21, 1, "EBITDA", bold=True)
for col in range(2, 17):
    cur_col = get_column_letter(col)
    formula_cell(ws, 21, col, f"={cur_col}12-{cur_col}19", CURR, bold=True)
total_row_format(ws, 21, 1, 16)

label_cell(ws, 22, 1, "  EBITDA margin %")
for col in range(2, 17):
    cur_col = get_column_letter(col)
    formula_cell(ws, 22, col, f"=IFERROR({cur_col}21/{cur_col}7,0)", PCT)

# Debt service
section_row(ws, 24, "Debt service & coverage", span=16)
label_cell(ws, 25, 1, "  Term debt service (P+I)")
# Approximate annual debt service: PMT-equivalent
# Total debt = $450K, 7-year, 6%. Annual service ~$80K/yr
for i in range(5):
    col = 2 + i
    formula_cell(ws, 25, col, f"=PMT({inp('Term debt interest rate')},{inp('Term debt amortization (years)')},-({inp('Term debt — Phase 1 portion')}+{inp('Term debt — Phase 2 portion')}))", CURR)
for i in range(5):
    formula_cell(ws, 25, 7 + i, f"={get_column_letter(2+i)}25", CURR)
    formula_cell(ws, 25, 12 + i, f"={get_column_letter(2+i)}25", CURR)

label_cell(ws, 26, 1, "  LoC interest")
for i in range(5):
    col = 2 + i
    if i == 0:
        formula_cell(ws, 26, col, f"={inp('Average LoC drawn (Y1)')}*{inp('LoC interest rate')}", CURR)
    else:
        formula_cell(ws, 26, col, f"={inp('Average LoC drawn (Y2)')}*{inp('LoC interest rate')}", CURR)
for i in range(5):
    formula_cell(ws, 26, 7 + i, f"={get_column_letter(2+i)}26", CURR)
    formula_cell(ws, 26, 12 + i, f"={get_column_letter(2+i)}26", CURR)

label_cell(ws, 27, 1, "Total debt service", bold=True)
for col in range(2, 17):
    cur_col = get_column_letter(col)
    formula_cell(ws, 27, col, f"={cur_col}25+{cur_col}26", CURR, bold=True)
total_row_format(ws, 27, 1, 16)

# DSCR
label_cell(ws, 28, 1, "DSCR (EBITDA / debt service)", bold=True)
for col in range(2, 17):
    cur_col = get_column_letter(col)
    formula_cell(ws, 28, col, f"=IFERROR({cur_col}21/{cur_col}27,0)", '0.00"x"', bold=True)
total_row_format(ws, 28, 1, 16)


# ===========================================================================
# WORKING CAPITAL
# ===========================================================================

ws = wb.create_sheet("Working Capital")
set_col_widths(ws, [38] + [14]*5)
header_row(ws, 1, ["Working Capital & Cash Conversion Cycle"] + years)

label_cell(ws, 3, 1, "Total revenue", bold=True)
for i in range(5):
    col = 2 + i
    formula_cell(ws, 3, col, f"='P&L Annual'!{get_column_letter(2+i)}7", CURR, bold=True)
    ws.cell(row=3, column=col).font = LINK_FONT

label_cell(ws, 4, 1, "Total COGS")
for i in range(5):
    col = 2 + i
    formula_cell(ws, 4, col, f"='P&L Annual'!{get_column_letter(2+i)}11", CURR)
    ws.cell(row=4, column=col).font = LINK_FONT

section_row(ws, 6, "Working capital balances (period-end)", span=6)

label_cell(ws, 7, 1, "Inventory (DIO * COGS / 365)")
for i in range(5):
    col = 2 + i
    cur_col = get_column_letter(col)
    dio_input = inp('Days Inventory Outstanding (Y1)') if i == 0 else inp('Days Inventory Outstanding (Y2+)')
    formula_cell(ws, 7, col, f"={cur_col}4*{dio_input}/365", CURR)

label_cell(ws, 8, 1, "Accounts payable (DPO * COGS / 365)")
for i in range(5):
    col = 2 + i
    cur_col = get_column_letter(col)
    dpo_input = inp('Days Payable Outstanding (Y1)') if i == 0 else inp('Days Payable Outstanding (Y2+)')
    formula_cell(ws, 8, col, f"={cur_col}4*{dpo_input}/365", CURR)

label_cell(ws, 9, 1, "Accounts receivable (DSO * revenue / 365)")
for i in range(5):
    col = 2 + i
    cur_col = get_column_letter(col)
    formula_cell(ws, 9, col, f"={cur_col}3*{inp('Days Sales Outstanding')}/365", CURR)

label_cell(ws, 10, 1, "Net working capital", bold=True)
for i in range(5):
    col = 2 + i
    cur_col = get_column_letter(col)
    formula_cell(ws, 10, col, f"={cur_col}7-{cur_col}8+{cur_col}9", CURR, bold=True)
total_row_format(ws, 10, 1, 6)

label_cell(ws, 12, 1, "Cash conversion cycle (DIO + DSO - DPO), days", bold=True)
for i in range(5):
    col = 2 + i
    dio = inp('Days Inventory Outstanding (Y1)') if i == 0 else inp('Days Inventory Outstanding (Y2+)')
    dpo = inp('Days Payable Outstanding (Y1)') if i == 0 else inp('Days Payable Outstanding (Y2+)')
    formula_cell(ws, 12, col, f"={dio}+{inp('Days Sales Outstanding')}-{dpo}", '0" days"', bold=True)
total_row_format(ws, 12, 1, 6)

label_cell(ws, 13, 1, "Change in NWC (cash impact)")
formula_cell(ws, 13, 2, "=B10-{}".format(inp('Opening assortment investment')), CURR)
ws.cell(row=13, column=2).font = LINK_FONT
for i in range(1, 5):
    col = 2 + i
    cur_col = get_column_letter(col)
    prev_col = get_column_letter(col - 1)
    formula_cell(ws, 13, col, f"={cur_col}10-{prev_col}10", CURR)

section_row(ws, 15, "Reserve and liquidity coverage", span=6)
label_cell(ws, 16, 1, "Operating reserve at close", bold=True)
formula_cell(ws, 16, 2, f"={inp('Operating reserve (subordinated)')}", CURR, bold=True)
label_cell(ws, 17, 1, "  Months of opex covered (Y2 stabilized)")
formula_cell(ws, 17, 2, f"=B16/('P&L Annual'!C19/12)", '0.0" months"')
label_cell(ws, 18, 1, "LoC facility size")
formula_cell(ws, 18, 2, f"={inp('Line of credit facility')}", CURR)
label_cell(ws, 19, 1, "  Days of COGS covered by LoC")
formula_cell(ws, 19, 2, "=B18/(C4/365)", '0" days"')


# ===========================================================================
# DEBT & DSCR
# ===========================================================================

ws = wb.create_sheet("Debt & DSCR")
set_col_widths(ws, [38] + [13]*5)
header_row(ws, 1, ["Debt Schedule & Coverage"] + years)

label_cell(ws, 3, 1, "Term debt — opening balance", bold=True)
formula_cell(ws, 3, 2, f"={inp('Term debt — Phase 1 portion')}+{inp('Term debt — Phase 2 portion')}", CURR, bold=True)
for i in range(1, 5):
    col = 2 + i
    prev_col = get_column_letter(col - 1)
    formula_cell(ws, 3, col, f"={prev_col}6", CURR, bold=True)

label_cell(ws, 4, 1, "  Annual debt service (P+I)")
for i in range(5):
    col = 2 + i
    formula_cell(ws, 4, col, f"=PMT({inp('Term debt interest rate')},{inp('Term debt amortization (years)')},-({inp('Term debt — Phase 1 portion')}+{inp('Term debt — Phase 2 portion')}))", CURR)

label_cell(ws, 5, 1, "  Of which: interest")
for i in range(5):
    col = 2 + i
    cur_col = get_column_letter(col)
    formula_cell(ws, 5, col, f"={cur_col}3*{inp('Term debt interest rate')}", CURR)

label_cell(ws, 6, 1, "Term debt — ending balance", bold=True)
for i in range(5):
    col = 2 + i
    cur_col = get_column_letter(col)
    formula_cell(ws, 6, col, f"=MAX(0,{cur_col}3-({cur_col}4-{cur_col}5))", CURR, bold=True)
total_row_format(ws, 6, 1, 6)

section_row(ws, 8, "Coverage ratios", span=6)
label_cell(ws, 9, 1, "EBITDA (base case)")
for i in range(5):
    col = 2 + i
    formula_cell(ws, 9, col, f"='P&L Annual'!{get_column_letter(2+i)}21", CURR)
    ws.cell(row=9, column=col).font = LINK_FONT

label_cell(ws, 10, 1, "Total debt service")
for i in range(5):
    col = 2 + i
    cur_col = get_column_letter(col)
    formula_cell(ws, 10, col, f"={cur_col}4+{inp('Average LoC drawn (Y1)')}*{inp('LoC interest rate')}", CURR)

label_cell(ws, 11, 1, "DSCR — base case", bold=True)
for i in range(5):
    col = 2 + i
    cur_col = get_column_letter(col)
    formula_cell(ws, 11, col, f"=IFERROR({cur_col}9/{cur_col}10,0)", '0.00"x"', bold=True)
total_row_format(ws, 11, 1, 6)

label_cell(ws, 12, 1, "DSCR — downside case")
for i in range(5):
    col = 2 + i
    cur_col = get_column_letter(col)
    formula_cell(ws, 12, col, f"=IFERROR('P&L Annual'!{get_column_letter(7+i)}21/{cur_col}10,0)", '0.00"x"')

label_cell(ws, 13, 1, "DSCR — stress case")
for i in range(5):
    col = 2 + i
    cur_col = get_column_letter(col)
    formula_cell(ws, 13, col, f"=IFERROR('P&L Annual'!{get_column_letter(12+i)}21/{cur_col}10,0)", '0.00"x"')

label_cell(ws, 15, 1, "DSCR target (CDFI standard)")
ws.cell(row=15, column=2, value="≥ 1.20x stabilized; 1.0x acceptable Y1 if reserves cover").font = Font(italic=True, color="666666")
ws.cell(row=15, column=2).alignment = LEFT
ws.merge_cells(start_row=15, start_column=2, end_row=15, end_column=6)


# ===========================================================================
# SENSITIVITY
# ===========================================================================

ws = wb.create_sheet("Sensitivity")
set_col_widths(ws, [40, 16, 16, 16, 16, 16])
header_row(ws, 1, ["Year 2 EBITDA Sensitivity", "-20%", "-10%", "Base", "+10%", "+20%"])

label_cell(ws, 3, 1, "Revenue (variable)")
y2_base_rev = "'P&L Annual'!C7"
y2_base_gm = "'P&L Annual'!C10"
y2_base_opex = "'P&L Annual'!C19"
for i, mult in enumerate([0.80, 0.90, 1.00, 1.10, 1.20]):
    col = 2 + i
    formula_cell(ws, 3, col, f"={y2_base_rev}*{mult}*{y2_base_gm}-{y2_base_opex}", CURR)

label_cell(ws, 4, 1, "Gross margin (variable, ±3 pts)")
for i, delta in enumerate([-0.03, -0.015, 0, 0.015, 0.03]):
    col = 2 + i
    formula_cell(ws, 4, col, f"={y2_base_rev}*({y2_base_gm}+{delta})-{y2_base_opex}", CURR)

label_cell(ws, 5, 1, "Labor as % of revenue (variable, ±2 pts)")
for i, delta in enumerate([0.02, 0.01, 0, -0.01, -0.02]):
    col = 2 + i
    # Labor up = EBITDA down
    formula_cell(ws, 5, col, f"={y2_base_rev}*{y2_base_gm}-{y2_base_opex}-{y2_base_rev}*{delta}", CURR)

section_row(ws, 7, "Two-way sensitivity: Revenue x Gross Margin", span=6)
header_row(ws, 8, ["Y2 EBITDA", "GM 26%", "GM 30%", "GM 34% (base)", "GM 38%", "GM 42%"])
gm_levels = [0.26, 0.30, 0.34, 0.38, 0.42]
rev_levels_lbl = ["Rev -20%", "Rev -10%", "Rev base", "Rev +10%", "Rev +20%"]
rev_mults = [0.80, 0.90, 1.00, 1.10, 1.20]
for i, lbl in enumerate(rev_levels_lbl):
    r = 9 + i
    label_cell(ws, r, 1, lbl)
    for j, gm in enumerate(gm_levels):
        col = 2 + j
        formula_cell(ws, r, col, f"={y2_base_rev}*{rev_mults[i]}*{gm}-{y2_base_opex}", CURR)

# Conditional formatting on the 2-way table
from openpyxl.styles import PatternFill as PF
from openpyxl.formatting.rule import CellIsRule
red_fill = PF(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
green_fill = PF(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ws.conditional_formatting.add("B9:F13", CellIsRule(operator="lessThan", formula=["0"], fill=red_fill))
ws.conditional_formatting.add("B9:F13", CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill))


# ===========================================================================
# MILESTONES
# ===========================================================================

ws = wb.create_sheet("Milestones")
set_col_widths(ws, [38, 18, 22, 60])
header_row(ws, 1, ["Tranche 2 release gate", "Threshold (end of M3)", "Verification method", "Why Steans cares"])

milestones = [
    ("Active produce-box subscribers", "100+ weekly subs", "Subscription management system export",
     "Demand validation, repeatable revenue, not survey-based intent"),
    ("Monthly revenue run-rate", "$15K - $25K / month", "Stripe / payment processor",
     "Operational viability; matches CDFI underwriting Year 1 ramp benchmark"),
    ("Local hires made", "5 - 7 NLW residents at $16+/hr", "Payroll records + zip code attestation",
     "Confirms workforce-development thesis; tied to grant deliverable"),
    ("EBT acceptance live", "Live by Month 2", "USDA SNAP retailer authorization letter",
     "Confirms Steans-priority access mechanism; required for low-income trade area"),
    ("Pembroke supply contracts", "Signed with 3+ farms", "Executed supply agreements",
     "Supply chain executable, not aspirational; differentiator from Yellow Banana"),
    ("Variance to budget", "Within 10% of plan", "Monthly P&L vs budget reconciliation",
     "Operator financial discipline"),
    ("Vendor terms earned", "Net 14 from 3+ DSD vendors", "Vendor account statements",
     "Working capital trajectory improving; reduces reliance on LoC"),
]

for i, (gate, threshold, method, why) in enumerate(milestones):
    r = 2 + i
    label_cell(ws, r, 1, gate)
    ws.cell(row=r, column=2, value=threshold).alignment = LEFT
    ws.cell(row=r, column=3, value=method).alignment = LEFT
    ws.cell(row=r, column=4, value=why).alignment = LEFT
    ws.row_dimensions[r].height = 35

section_row(ws, 12, "Tranche 2 release decision matrix", span=4)
ws.cell(row=13, column=1, value="If 7 of 7 gates pass:").font = TOTAL_FONT
ws.cell(row=13, column=2, value="Full Tranche 2 release ($500K)").alignment = LEFT
ws.cell(row=14, column=1, value="If 5-6 of 7 gates pass:").font = TOTAL_FONT
ws.cell(row=14, column=2, value="Conditional release with 30-day cure period for missed gates").alignment = LEFT
ws.cell(row=15, column=1, value="If <5 gates pass:").font = TOTAL_FONT
ws.cell(row=15, column=2, value="Tranche 2 deferred; operator and funder reassess plan").alignment = LEFT


# ===========================================================================
# IMPACT METRICS
# ===========================================================================

ws = wb.create_sheet("Impact")
set_col_widths(ws, [42] + [14]*5)
header_row(ws, 1, ["Steans Impact Metric"] + years)

label_cell(ws, 2, 1, "WORKFORCE", bold=True)
section_row(ws, 2, "WORKFORCE", span=6)

label_cell(ws, 3, 1, "  FTE — total")
y1_avg_fte = f"({inp('Phase 1 FTE — Months 1-3')}*3+{inp('Phase 1 FTE — Months 4-6')}*3+{inp('Phase 2 FTE — Months 7-9')}*3+{inp('Phase 2 FTE — Months 10-12')}*3)/12"
formula_cell(ws, 3, 2, f"={y1_avg_fte}", '0.0" FTE"')
formula_cell(ws, 3, 3, f"={inp('Y2 FTE — average')}", '0.0" FTE"')
formula_cell(ws, 3, 4, f"={inp('Y3 FTE — average')}", '0.0" FTE"')
formula_cell(ws, 3, 5, f"={inp('Y4 FTE — average')}", '0.0" FTE"')
formula_cell(ws, 3, 6, f"={inp('Y5 FTE — average')}", '0.0" FTE"')

label_cell(ws, 4, 1, "  % hires from NLW zip codes (target)")
for col in range(2, 7):
    formula_cell(ws, 4, col, "=0.7", PCT_INT)

label_cell(ws, 5, 1, "  Hourly wage floor")
for col in range(2, 7):
    formula_cell(ws, 5, col, "=16", '"$"0.00"/hr"')

label_cell(ws, 6, 1, "  Total annual payroll (NLW residents)")
formula_cell(ws, 6, 2, "=B3*B4*16*173*12", CURR)
for i in range(1, 5):
    col = 2 + i
    cur_col = get_column_letter(col)
    formula_cell(ws, 6, col, f"={cur_col}3*{cur_col}4*16*173*12", CURR)

section_row(ws, 8, "FOOD ACCESS", span=6)
label_cell(ws, 9, 1, "  Households served / month (subs + retail)")
# Y1: subscribers + retail transactions / 4 (assume 4 visits/mo per HH retail)
formula_cell(ws, 9, 2, f"='Phase 1 Revenue'!M4+ROUND('P&L Annual'!B6/12/45,0)/4", NUM)  # avg basket $45
for i in range(1, 5):
    col = 2 + i
    cur_col = get_column_letter(col)
    sub_lbl = ["Y2 produce subscribers (avg active)", "Y3 produce subscribers (avg active)",
               "Y4 produce subscribers (avg active)", "Y5 produce subscribers (avg active)"][i-1]
    formula_cell(ws, 9, col, f"={inp(sub_lbl)}+ROUND('P&L Annual'!{get_column_letter(2+i)}6/12/45,0)/4", NUM)

label_cell(ws, 10, 1, "  SNAP/EBT as % of retail sales (target)")
for col in range(3, 7):
    formula_cell(ws, 10, col, "=0.20", PCT_INT)

label_cell(ws, 11, 1, "  USDA Low Access tract: site at 17031843300")
ws.cell(row=11, column=2, value="LILA designated").alignment = LEFT
ws.merge_cells(start_row=11, start_column=2, end_row=11, end_column=6)

section_row(ws, 13, "SUPPLIER NETWORK", span=6)
label_cell(ws, 14, 1, "  Pembroke direct procurement ($)")
for i in range(5):
    col = 2 + i
    cur_col = get_column_letter(col)
    # Approx 30% of produce spend
    formula_cell(ws, 14, col, f"='P&L Annual'!{cur_col}11*0.18*0.3", CURR)  # 18% mix is produce, 30% from Pembroke

label_cell(ws, 15, 1, "  Black-owned vendor procurement ($)")
for i in range(5):
    col = 2 + i
    cur_col = get_column_letter(col)
    formula_cell(ws, 15, col, f"='P&L Annual'!{cur_col}11*0.25", CURR)  # 25% of total COGS to Black-owned

section_row(ws, 17, "LEVERAGE", span=6)
label_cell(ws, 18, 1, "  Steans capital deployed (cumulative)")
formula_cell(ws, 18, 2, f"={inp('Tranche 1 (Phase 1 launch)')}", CURR)
for i in range(1, 5):
    col = 2 + i
    formula_cell(ws, 18, col, f"={inp('Tranche 1 (Phase 1 launch)')}+{inp('Tranche 2 (Phase 2 retail)')}", CURR)

label_cell(ws, 19, 1, "  Total capital deployed (with LoC drawn)")
formula_cell(ws, 19, 2, f"={inp('Tranche 1 (Phase 1 launch)')}+{inp('Average LoC drawn (Y1)')}", CURR)
for i in range(1, 5):
    col = 2 + i
    formula_cell(ws, 19, col, f"={inp('Tranche 1 (Phase 1 launch)')}+{inp('Tranche 2 (Phase 2 retail)')}+{inp('Average LoC drawn (Y2)')}", CURR)

label_cell(ws, 20, 1, "  Cumulative revenue generated")
formula_cell(ws, 20, 2, "='P&L Annual'!B7", CURR)
for i in range(1, 5):
    col = 2 + i
    cur_col = get_column_letter(col)
    prev_col = get_column_letter(col - 1)
    formula_cell(ws, 20, col, f"={prev_col}20+'P&L Annual'!{cur_col}7", CURR)

label_cell(ws, 21, 1, "  Revenue per dollar of Steans capital (cumulative)", bold=True)
for col in range(2, 7):
    cur_col = get_column_letter(col)
    formula_cell(ws, 21, col, f"={cur_col}20/{cur_col}18", '0.00"x"', bold=True)
total_row_format(ws, 21, 1, 6)


# ===========================================================================
# Reorder tabs and save
# ===========================================================================

order = ["README", "Inputs", "Capital Stack", "Phase 1 Revenue", "Phase 2 Revenue",
         "P&L Monthly Y1", "P&L Annual", "Working Capital", "Debt & DSCR",
         "Sensitivity", "Milestones", "Impact"]
wb._sheets = [wb[s] for s in order]

# Set tab colors
wb["README"].sheet_properties.tabColor = "1F4E2C"
wb["Inputs"].sheet_properties.tabColor = "1F4E2C"
wb["Capital Stack"].sheet_properties.tabColor = "C9A02E"
wb["Phase 1 Revenue"].sheet_properties.tabColor = "C9A02E"
wb["Phase 2 Revenue"].sheet_properties.tabColor = "C9A02E"
wb["P&L Monthly Y1"].sheet_properties.tabColor = "555555"
wb["P&L Annual"].sheet_properties.tabColor = "555555"
wb["Working Capital"].sheet_properties.tabColor = "555555"
wb["Debt & DSCR"].sheet_properties.tabColor = "555555"
wb["Sensitivity"].sheet_properties.tabColor = "888888"
wb["Milestones"].sheet_properties.tabColor = "888888"
wb["Impact"].sheet_properties.tabColor = "888888"

# Freeze panes on key sheets
for sheet in ["Inputs", "Capital Stack"]:
    wb[sheet].freeze_panes = "A2"
for sheet in ["Phase 1 Revenue", "Phase 2 Revenue", "P&L Monthly Y1", "P&L Annual",
              "Working Capital", "Debt & DSCR", "Impact"]:
    wb[sheet].freeze_panes = "B3"

out = "uncle-mays-financial-model.xlsx"
wb.save(out)
print(f"Saved {out}")
